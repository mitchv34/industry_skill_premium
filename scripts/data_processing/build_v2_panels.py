"""Build v2 per-industry panels with industry-specific equipment deflators.

Source: BEA Fixed Assets Section 3, sheets FAAt307E-A (nominal equipment
investment by industry, $M, 1947-2024) and FAAt308E-A (chain-type quantity
index for equipment investment, 2017=100).

For each industry, the implied investment-price deflator is:

    P_{i,t} = (nominal investment_{i,t}) / (quantity index_{i,t})

This is a Tornqvist-aggregated price index, already industry-specific, computed
internally by BEA — strictly cleaner than rebuilding the Tornqvist from
asset×industry detail. See `data/raw/alternative_sources/README.md` for the
methodology rationale.

Output: `data/proc/ind_v2/{IND}.csv` for each of the 56 project industries.
The original `data/proc/ind/{IND}.csv` panels are preserved; v2 panels replace
`REL_P_EQ` with the industry-specific deflator and store the old aggregate
series as `REL_P_EQ_AGG` for robustness.

Run:
    .venv/bin/python scripts/data_processing/build_v2_panels.py
"""

from __future__ import annotations

import re
import sys
import warnings
from pathlib import Path

import polars as pl
import openpyxl

warnings.simplefilter("ignore")

# Repo paths
ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(ROOT))  # so we can `from config import ...` and from build_industry_deflators
import config  # noqa: E402
from scripts.data_processing.build_industry_deflators import (  # noqa: E402
    merge_into_industry_panel,
)

SECTION3 = ROOT / "data" / "raw" / "alternative_sources" / "bea_fixed_assets" / "Section3All.xlsx"
INDUSTRY_NAMES = ROOT / "data" / "industry_names.csv"
PANEL_V1 = ROOT / "data" / "proc" / "ind"
PANEL_V2 = ROOT / "data" / "proc" / "ind_v2"


# ---------------------------------------------------------------------------
# 1. BEA Section 3 sheet parser
# ---------------------------------------------------------------------------

def parse_section3_sheet(workbook: openpyxl.Workbook, sheet_name: str) -> pl.DataFrame:
    """Parse a Section3All sheet (e.g. FAAt307E-A) into long format.

    Returns DataFrame with schema:
        INDUSTRY_NAME (str) — col 2, trimmed and footnote-stripped
        SERIES_CODE   (str) — col 3, the BEA series identifier
        YEAR          (int)
        VALUE         (float)

    The sheet layout (verified for Section3All as released Sept 2025):
        row 1-7: metadata / blank
        row 8:   header — "Line", "", "", 1947, 1948, ..., 2024
        row 9+:  data — line#, industry name, series code, val_1947, val_1948, ...
    """
    ws = workbook[sheet_name]
    # Single linear pass via iter_rows — openpyxl's read-only mode is slow
    # under ws.cell(r,c) random access but fast for sequential row iteration.

    year_cols: list[tuple[int, int]] = []
    rows: list[dict] = []
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if r_idx == 8:
            # Header row: discover year columns.
            for c_idx, v in enumerate(row, start=1):
                if v is None:
                    continue
                try:
                    year = int(str(v).strip())
                except (TypeError, ValueError):
                    continue
                if 1940 < year < 2030:
                    year_cols.append((c_idx, year))
            if not year_cols:
                raise ValueError(
                    f"Sheet {sheet_name}: no year columns in row 8 (saw: {row[:10]!r})"
                )
        elif r_idx >= 9:
            # Data row: col 2 = industry name, col 3 = series code, year_cols = values.
            if len(row) < 3:
                continue
            name_raw = row[1]
            code = row[2]
            if name_raw is None or code is None:
                continue
            name = _normalize_industry_name(str(name_raw))
            code_s = str(code).strip()
            for c_idx, year in year_cols:
                v = row[c_idx - 1] if c_idx - 1 < len(row) else None
                if v is None or v == "":
                    continue
                try:
                    val = float(v)
                except (TypeError, ValueError):
                    continue
                rows.append({
                    "INDUSTRY_NAME": name,
                    "SERIES_CODE": code_s,
                    "YEAR": year,
                    "VALUE": val,
                })

    return pl.DataFrame(rows)


def _normalize_industry_name(s: str) -> str:
    """Strip BEA footnote markers like ' 1', ' \\1\\', trailing whitespace.

    Examples:
        'Farms 1'             -> 'Farms'
        '  Farms \\1\\'        -> 'Farms'
        'Retail trade'        -> 'Retail trade'
    """
    s = s.strip()
    # Strip backslash-delimited footnote refs: \1\, \2\, etc.
    s = re.sub(r"\\\d+\\", "", s).strip()
    # Strip trailing footnote integer like ' 1' or ' 2'
    s = re.sub(r"\s+\d+$", "", s).strip()
    return s


# ---------------------------------------------------------------------------
# 2. Industry crosswalk: BEA industry name -> project BEA_Code
# ---------------------------------------------------------------------------

# Name overrides where the project's display name differs from BEA's.
BEA_NAME_OVERRIDES: dict[str, str] = {
    "Rail transportation": "Railroad transportation",
}

# Aggregations where one project industry is composed of multiple BEA industries.
# These are aggregated via nominal-investment-weighted average of implied prices.
BEA_AGGREGATIONS: dict[str, list[str]] = {
    "521CI": [
        "Federal Reserve banks",
        "Credit intermediation and related activities",
    ],
    "622HO": [
        "Hospitals",
        "Nursing and residential care facilities",
    ],
}


def build_industry_crosswalk() -> pl.DataFrame:
    """Map BEA industry display names to the project's 56 BEA_Code values.

    The project's `data/industry_names.csv` already uses BEA's industry names
    verbatim with a few exceptions. Exceptions are handled via:
    - BEA_NAME_OVERRIDES — direct rename (e.g., 'Rail' → 'Railroad').
    - BEA_AGGREGATIONS — sum/avg over BEA sub-industries to match a project
      composite (e.g., '622HO' = 'Hospitals' + 'Nursing and residential care').
      These are handled separately in `compute_industry_deflator`.

    Returns the simple name->code mapping only; aggregations are layered on top.
    """
    proj = pl.read_csv(INDUSTRY_NAMES)
    return proj.with_columns(
        INDUSTRY_NAME=pl.col("Description").map_elements(
            lambda s: BEA_NAME_OVERRIDES.get(_normalize_industry_name(s), _normalize_industry_name(s)),
            return_dtype=pl.Utf8,
        ),
        BEA_CODE=pl.col("BEA_Code"),
    ).select(["INDUSTRY_NAME", "BEA_CODE"])


# ---------------------------------------------------------------------------
# 3. Build the industry-specific equipment deflator from FAAt307E + FAAt308E.
# ---------------------------------------------------------------------------

def compute_industry_deflator(workbook: openpyxl.Workbook) -> pl.DataFrame:
    """Compute the industry-specific equipment-investment price *deviation*.

    The v1 panels use `REL_P_EQ` = the aggregate relative price of equipment to
    consumption (KORV's q_t, FRED PERIC-style), falling from 1.0 (1987) to ~0.13
    (2018). This is *the relative price KORV uses in the model*.

    My ratio T3.7E_i / T3.8E_i gives the nominal equipment investment-price
    index per industry — NOT relative to consumption. To make the industry
    series comparable to the v1 aggregate (model-consistent), I compute the
    industry's deviation from the aggregate and multiply by the v1 aggregate:

        ratio_{i,t}  = (P_industry^E_{i,t} / P_aggregate^E_{t})
                       normalized so that ratio_{i,1987} = 1.0
        q_{i,t}      = ratio_{i,t} × q_aggregate_{t}    (= ratio × v1 REL_P_EQ)

    This preserves the KORV interpretation (relative price of equipment to
    consumption) while picking up industry heterogeneity in equipment
    composition. Industries with cheaper-than-average equipment baskets (e.g.,
    legal services with lots of IT) see q_{i,t} fall *faster* than the
    aggregate; industries with non-IT-heavy baskets see q_{i,t} fall *slower*.

    The multiplication by v1 REL_P_EQ_AGG is done downstream in the merge step
    (see main()), so this function just returns the industry deviation.

    Returns DataFrame with schema:
        BEA_CODE (str)
        YEAR     (int)
        REL_P_EQ_DEVIATION (float) — industry equipment price / aggregate
            equipment price, normalized to 1.0 in 1987.
    """
    inv = parse_section3_sheet(workbook, "FAAt307E-A").rename({"VALUE": "INVESTMENT_NOMINAL"})
    qty = parse_section3_sheet(workbook, "FAAt308E-A").rename({"VALUE": "QUANTITY_INDEX"})

    # Join nominal × quantity by (industry name, year). Drop the SERIES_CODE.
    base = (
        inv.drop("SERIES_CODE")
        .join(qty.drop("SERIES_CODE"), on=["INDUSTRY_NAME", "YEAR"], how="inner")
        .with_columns(
            implied_p=pl.col("INVESTMENT_NOMINAL") / pl.col("QUANTITY_INDEX")
        )
    )

    # Direct name matches via the crosswalk.
    cw = build_industry_crosswalk()
    deflator_direct = base.join(cw, on="INDUSTRY_NAME", how="inner").select([
        "BEA_CODE", "YEAR", "INVESTMENT_NOMINAL", "QUANTITY_INDEX", "implied_p"
    ])

    # Composite industries: investment-weighted average of implied prices.
    composite_rows = []
    for proj_code, bea_names in BEA_AGGREGATIONS.items():
        comp = base.filter(pl.col("INDUSTRY_NAME").is_in(bea_names))
        if comp.is_empty():
            print(f"  warn: aggregation {proj_code} matched zero rows in BEA data")
            continue
        agg = (
            comp.group_by("YEAR")
            .agg(
                INVESTMENT_NOMINAL=pl.col("INVESTMENT_NOMINAL").sum(),
                implied_p_weighted=(pl.col("implied_p") * pl.col("INVESTMENT_NOMINAL")).sum()
                                   / pl.col("INVESTMENT_NOMINAL").sum(),
            )
            .with_columns(
                BEA_CODE=pl.lit(proj_code),
                QUANTITY_INDEX=pl.lit(None, dtype=pl.Float64),  # not meaningful for composite
                implied_p=pl.col("implied_p_weighted"),
            )
            .select(["BEA_CODE", "YEAR", "INVESTMENT_NOMINAL", "QUANTITY_INDEX", "implied_p"])
        )
        composite_rows.append(agg)

    if composite_rows:
        deflator = pl.concat([deflator_direct] + composite_rows, how="vertical")
    else:
        deflator = deflator_direct

    # Build the aggregate "Private fixed assets" deflator from base.
    # This is BEA's official aggregate equipment price index across all industries.
    agg = (
        base.filter(pl.col("INDUSTRY_NAME") == "Private fixed assets")
        .select(["YEAR", "implied_p"])
        .rename({"implied_p": "implied_p_agg"})
    )
    if agg.is_empty():
        raise RuntimeError("Could not find aggregate 'Private fixed assets' row in BEA data")

    # Compute industry deviation = P_industry / P_aggregate, normalized to 1.0 in 1987.
    deflator = deflator.join(agg, on="YEAR", how="left").with_columns(
        ratio=pl.col("implied_p") / pl.col("implied_p_agg")
    )
    base_year = 1987
    base_ratio = (
        deflator.filter(pl.col("YEAR") == base_year)
        .select(["BEA_CODE", "ratio"])
        .rename({"ratio": "_base_ratio"})
    )
    deflator = deflator.join(base_ratio, on="BEA_CODE", how="left").with_columns(
        REL_P_EQ_DEVIATION=pl.col("ratio") / pl.col("_base_ratio")
    )

    return deflator.select(["BEA_CODE", "YEAR", "REL_P_EQ_DEVIATION"])


# ---------------------------------------------------------------------------
# 4. Main: assemble v2 panels.
# ---------------------------------------------------------------------------

def main() -> None:
    print(f"Loading Section3All workbook: {SECTION3}")
    wb = openpyxl.load_workbook(SECTION3, data_only=True, read_only=True)

    print("Computing industry-specific equipment deflator deviations (1947–2024) ...")
    deviation = compute_industry_deflator(wb)
    print(f"  -> {deviation.height} industry-year rows")
    print(f"  -> {deviation['BEA_CODE'].n_unique()} unique BEA_CODEs covered")
    print(f"  -> year range: {deviation['YEAR'].min()}–{deviation['YEAR'].max()}")

    # Sanity: NAICS 334 (computers) should have deviation falling well below 1
    # by 2010 (IT cheaper than average equipment); 23 (construction) should
    # have deviation > 1 (non-IT equipment more expensive than avg over time).
    chk = (
        deviation.filter(pl.col("YEAR") == 2010)
        .select(["BEA_CODE", "REL_P_EQ_DEVIATION"])
    )
    for code in ["334", "23", "5411", "311FT", "55", "5415"]:
        row = chk.filter(pl.col("BEA_CODE") == code)
        if row.height:
            d = row["REL_P_EQ_DEVIATION"][0]
            note = ""
            if code == "334":
                note = " ← computers, expect < 1"
            elif code == "5411":
                note = " ← legal services, expect << 1 (IT-heavy)"
            elif code == "23":
                note = " ← construction, expect > 1 (non-IT)"
            print(f"  sanity {code}: deviation_2010 = {d:.3f}{note}")

    # Save deviation as a standalone reference file.
    deviation_csv = ROOT / "data" / "proc" / "industry_relative_prices.csv"
    deviation.write_csv(deviation_csv)
    print(f"  wrote {deviation_csv}")

    # Merge into each v1 panel.
    # NEW REL_P_EQ = DEVIATION × old REL_P_EQ (the aggregate KORV q_t).
    PANEL_V2.mkdir(parents=True, exist_ok=True)
    v1_files = sorted(PANEL_V1.glob("*.csv"))
    print(f"\nMerging into {len(v1_files)} v1 panels at {PANEL_V1} -> {PANEL_V2}")

    successes, misses = [], []
    for v1 in v1_files:
        bea_code = v1.stem
        panel = pl.read_csv(v1)
        ind_dev = deviation.filter(pl.col("BEA_CODE") == bea_code).select(
            ["YEAR", "REL_P_EQ_DEVIATION"]
        )
        if ind_dev.height == 0:
            misses.append(bea_code)
            continue

        # Preserve the old aggregate series.
        merged = panel.rename({"REL_P_EQ": "REL_P_EQ_AGG"})

        # Multiply industry deviation × aggregate to get industry-specific q_t.
        merged = (
            merged.join(ind_dev, on="YEAR", how="left")
            .with_columns(
                REL_P_EQ=pl.col("REL_P_EQ_DEVIATION") * pl.col("REL_P_EQ_AGG"),
            )
            .drop("REL_P_EQ_DEVIATION")
        )

        # If deviation missing for any year, fall back to AGG (no-deviation).
        merged = merged.with_columns(
            REL_P_EQ=pl.coalesce(["REL_P_EQ", "REL_P_EQ_AGG"])
        )

        # Restore original column order, with AGG appended.
        new_order = list(panel.columns) + ["REL_P_EQ_AGG"]
        merged = merged.select(new_order)

        out = PANEL_V2 / v1.name
        merged.write_csv(out)
        successes.append(bea_code)

    print(f"\nWrote {len(successes)} v2 panels.")
    if misses:
        print(f"Industries with no BEA deflator (kept v1 as-is, not copied): {misses}")
    else:
        print("All v1 panels successfully merged.")


if __name__ == "__main__":
    main()
