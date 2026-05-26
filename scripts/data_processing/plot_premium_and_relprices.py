"""Two figures for Mitchell, v2:

  fig1: skill premium 1963–2025 — aggregate + 56 industry-specific overlays,
        normalized to 1.0 at the earliest year with broad industry coverage.
  fig2: industry equipment relative price 1947–2024, all 56 industries,
        normalized to 1.0 at the same base year as fig1.

Pre-1976 CPS lacks UHRSWORKLY / WKSWORK1, so we impute:
  - WKSWORK1 ← midpoint of WKSWORK2 bins
  - UHRSWORKLY ← AHRSWORKT (hours worked last week) as a proxy

Industry mapping uses `data/cross_walk.csv` (Census IND1990 → BEA_Code).

Run:
    .venv/bin/python scripts/data_processing/plot_premium_and_relprices.py
"""
from __future__ import annotations

from pathlib import Path
import sys
import warnings

import matplotlib.pyplot as plt
import openpyxl
import polars as pl

warnings.simplefilter("ignore")

ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(ROOT))

CPS = ROOT / "data" / "raw" / "cps" / "cps_extract_106.csv.gz"
SECTION3 = ROOT / "data" / "raw" / "alternative_sources" / "bea_fixed_assets" / "Section3All.xlsx"
CROSSWALK = ROOT / "data" / "cross_walk.csv"
INDUSTRY_NAMES = ROOT / "data" / "industry_names.csv"
OUT = ROOT / "documents" / "images"
OUT.mkdir(parents=True, exist_ok=True)

# Reused from build_v2_panels — same name overrides + aggregations.
BEA_NAME_OVERRIDES = {"Rail transportation": "Railroad transportation"}
BEA_AGGREGATIONS = {
    "521CI": ["Federal Reserve banks", "Credit intermediation and related activities"],
    "622HO": ["Hospitals", "Nursing and residential care facilities"],
}


def style_axes(ax: plt.Axes) -> None:
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_visible(False)
    ax.grid(False)
    ax.tick_params(left=False, bottom=True)


# ===========================================================================
# 1. CPS — aggregate + industry-specific skill premium
# ===========================================================================

def _wkswork2_midpoint() -> pl.Expr:
    """Hot-deck-free midpoint for the WKSWORK2 banded weeks variable.

    CPS WKSWORK2 codes:
        0 = N/A   1 = 1–13   2 = 14–26   3 = 27–39
        4 = 40–47 5 = 48–49  6 = 50–52
    """
    return (
        pl.when(pl.col("WKSWORK2") == 1).then(7.0)
          .when(pl.col("WKSWORK2") == 2).then(20.0)
          .when(pl.col("WKSWORK2") == 3).then(33.0)
          .when(pl.col("WKSWORK2") == 4).then(43.5)
          .when(pl.col("WKSWORK2") == 5).then(48.5)
          .when(pl.col("WKSWORK2") == 6).then(51.0)
          .otherwise(0.0)
    )


def _load_cps_filtered() -> pl.DataFrame:
    """Load and filter CPS extract, applying pre-1976 hours imputation."""
    print(f"Loading {CPS} (205 MB)...")
    df = pl.read_csv(CPS, infer_schema_length=10000, null_values=["", "NA", "."])
    print(f"  raw rows: {df.height:,}")

    for c in ["WKSWORK1", "UHRSWORKLY", "OINCWAGE", "EDUC99"]:
        if c in df.columns:
            df = df.with_columns(pl.col(c).cast(pl.Float64, strict=False))

    # Impute weeks & hours pre-1976 (CPS didn't collect WKSWORK1/UHRSWORKLY).
    df = df.with_columns(
        weeks_imp=pl.coalesce([pl.col("WKSWORK1"), _wkswork2_midpoint()]),
        hours_imp=pl.coalesce([pl.col("UHRSWORKLY"), pl.col("AHRSWORKT").cast(pl.Float64)]),
    )

    df = df.filter(
        pl.col("ASECWT") > 0,
        pl.col("AGE").is_between(16, 70),
        pl.col("EMPSTAT").is_in([10, 12]),
        pl.col("CLASSWLY").is_in([13, 14, 21, 22, 24, 28]),
        pl.col("WKSWORK2") >= 5,                  # 40+ weeks
        pl.col("hours_imp") >= 30,                # 30+ hrs/week
        pl.col("INCWAGE") > 0,
        pl.col("INCWAGE") < 99999998,
        pl.col("OINCWAGE").is_null() | (pl.col("OINCWAGE") == 0),
    )
    print(f"  after filters: {df.height:,}")

    df = df.with_columns(
        hourly_wage=pl.col("INCWAGE") / (pl.col("weeks_imp") * pl.col("hours_imp"))
    ).filter(pl.col("hourly_wage") > 0)

    df = df.with_columns(
        skill=pl.when(pl.col("EDUC") >= 80).then(pl.lit("college")).otherwise(pl.lit("noncollege")),
        IND1990=pl.col("IND1990").cast(pl.Int64, strict=False),
    )
    return df


def build_skill_premium(df_cps: pl.DataFrame) -> pl.DataFrame:
    """Aggregate skill premium per year, 1963–2025."""
    agg = (
        df_cps.group_by(["YEAR", "skill"])
        .agg(
            mean_wage=(pl.col("hourly_wage") * pl.col("ASECWT")).sum() / pl.col("ASECWT").sum(),
            n=pl.len(),
        )
        .pivot(values="mean_wage", index="YEAR", on="skill")
        .sort("YEAR")
        .with_columns(SKILL_PREMIUM=pl.col("college") / pl.col("noncollege"))
        .rename({"college": "W_college", "noncollege": "W_noncollege"})
    )
    return agg


def _ind_crosswalk() -> dict[int, str]:
    """Build IND1990 (Census) → project BEA_Code mapping from cross_walk.csv."""
    cw = pl.read_csv(CROSSWALK)
    mapping: dict[int, str] = {}
    for row in cw.iter_rows(named=True):
        bea = row["code_bea"]
        # Project codes in industry_names.csv come from `code_klems` not code_bea,
        # but the cross_walk file lists code_klems too — use it for compatibility.
        bea_proj = str(row["code_klems"])
        census_csv = str(row["code_census"]) if row["code_census"] is not None else ""
        for c in census_csv.split(","):
            c = c.strip()
            if c.isdigit():
                mapping[int(c)] = bea_proj
    return mapping


def build_industry_skill_premium(df_cps: pl.DataFrame, min_n: int = 50) -> pl.DataFrame:
    """Skill premium per (YEAR, BEA_CODE), requiring >= min_n obs per cell."""
    ind_map = _ind_crosswalk()
    df = df_cps.with_columns(
        BEA_CODE=pl.col("IND1990").replace_strict(ind_map, default=None, return_dtype=pl.Utf8)
    ).filter(pl.col("BEA_CODE").is_not_null())

    agg = (
        df.group_by(["YEAR", "BEA_CODE", "skill"])
        .agg(
            mean_wage=(pl.col("hourly_wage") * pl.col("ASECWT")).sum() / pl.col("ASECWT").sum(),
            n=pl.len(),
        )
        .pivot(values=["mean_wage", "n"], index=["YEAR", "BEA_CODE"], on="skill")
    )
    # After pivot, columns are mean_wage_college, mean_wage_noncollege, n_college, n_noncollege
    # Filter for cells with enough observations on BOTH sides.
    agg = agg.filter(
        (pl.col("n_college") >= min_n) & (pl.col("n_noncollege") >= min_n)
    ).with_columns(
        SKILL_PREMIUM=pl.col("mean_wage_college") / pl.col("mean_wage_noncollege")
    ).select(["YEAR", "BEA_CODE", "SKILL_PREMIUM"])

    return agg.sort(["BEA_CODE", "YEAR"])


# ===========================================================================
# 2. BEA — full-range industry equipment prices, normalized at base_year
# ===========================================================================

def _normalize_industry_name(s: str) -> str:
    import re
    s = s.strip()
    s = re.sub(r"\\\d+\\", "", s).strip()
    s = re.sub(r"\s+\d+$", "", s).strip()
    return s


def _parse_section3_long(wb, sheet_name) -> pl.DataFrame:
    ws = wb[sheet_name]
    year_cols: list[tuple[int, int]] = []
    rows = []
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if r_idx == 8:
            for c_idx, v in enumerate(row, start=1):
                try:
                    yr = int(str(v).strip())
                    if 1940 < yr < 2030:
                        year_cols.append((c_idx, yr))
                except (TypeError, ValueError):
                    pass
        elif r_idx >= 9:
            if len(row) < 3 or row[1] is None or row[2] is None:
                continue
            name = _normalize_industry_name(str(row[1]))
            for c_idx, year in year_cols:
                if c_idx - 1 >= len(row):
                    continue
                v = row[c_idx - 1]
                if v is None or v == "":
                    continue
                try:
                    val = float(v)
                except (TypeError, ValueError):
                    continue
                rows.append({"INDUSTRY_NAME": name, "YEAR": year, "VALUE": val})
    return pl.DataFrame(rows)


def build_industry_eq_prices(base_year: int) -> tuple[pl.DataFrame, pl.DataFrame]:
    """Full 1947–2024 absolute equipment price index per industry + aggregate.

    Returns (industry_long, aggregate_long), both normalized to 1.0 at base_year.

    industry_long schema: [BEA_CODE, YEAR, P_norm]
    aggregate_long schema: [YEAR, P_agg_norm]
    """
    print(f"Building BEA equipment prices, normalized to {base_year}=1.0 ...")
    wb = openpyxl.load_workbook(SECTION3, data_only=True, read_only=True)
    inv = _parse_section3_long(wb, "FAAt307E-A").rename({"VALUE": "INV"})
    qty = _parse_section3_long(wb, "FAAt308E-A").rename({"VALUE": "QTY"})
    base = inv.join(qty, on=["INDUSTRY_NAME", "YEAR"]).with_columns(P=pl.col("INV") / pl.col("QTY"))

    # Aggregate: "Private fixed assets" line.
    agg = base.filter(pl.col("INDUSTRY_NAME") == "Private fixed assets").select(["YEAR", "P"]).rename({"P": "P_agg"})
    p_agg_base = agg.filter(pl.col("YEAR") == base_year)["P_agg"][0]
    aggregate_long = agg.with_columns(P_agg_norm=pl.col("P_agg") / p_agg_base).select(["YEAR", "P_agg_norm"])

    # Industry direct matches via name → BEA_Code crosswalk.
    proj = pl.read_csv(INDUSTRY_NAMES)
    proj_map = {
        BEA_NAME_OVERRIDES.get(_normalize_industry_name(d), _normalize_industry_name(d)): c
        for d, c in zip(proj["Description"].to_list(), proj["BEA_Code"].to_list())
    }
    direct = base.with_columns(
        BEA_CODE=pl.col("INDUSTRY_NAME").replace_strict(proj_map, default=None, return_dtype=pl.Utf8)
    ).filter(pl.col("BEA_CODE").is_not_null()).select(["BEA_CODE", "YEAR", "P"])

    # Composite aggregations (521CI, 622HO): investment-weighted average price.
    comp_rows = []
    for proj_code, bea_names in BEA_AGGREGATIONS.items():
        comp = base.filter(pl.col("INDUSTRY_NAME").is_in(bea_names))
        if comp.is_empty():
            continue
        a = (
            comp.group_by("YEAR")
            .agg(
                INV_total=pl.col("INV").sum(),
                P=(pl.col("P") * pl.col("INV")).sum() / pl.col("INV").sum(),
            )
            .with_columns(BEA_CODE=pl.lit(proj_code))
            .select(["BEA_CODE", "YEAR", "P"])
        )
        comp_rows.append(a)
    industry_p = pl.concat([direct] + comp_rows, how="vertical") if comp_rows else direct

    # Normalize each industry to 1.0 at base_year.
    bases = (
        industry_p.filter(pl.col("YEAR") == base_year)
        .select(["BEA_CODE", "P"])
        .rename({"P": "_base"})
    )
    industry_long = industry_p.join(bases, on="BEA_CODE").with_columns(
        P_norm=pl.col("P") / pl.col("_base")
    ).select(["BEA_CODE", "YEAR", "P_norm"])

    return industry_long, aggregate_long


# ===========================================================================
# Plotting
# ===========================================================================

def _normalize_series(df: pl.DataFrame, value_col: str, group_col: str | None, base_year: int) -> pl.DataFrame:
    """Normalize value_col so it equals 1.0 at base_year. Optionally per-group."""
    if group_col:
        bases = (
            df.filter(pl.col("YEAR") == base_year)
            .select([group_col, value_col]).rename({value_col: "_base"})
        )
        return df.join(bases, on=group_col).with_columns(
            (pl.col(value_col) / pl.col("_base")).alias(value_col)
        ).drop("_base")
    else:
        base = df.filter(pl.col("YEAR") == base_year)[value_col][0]
        return df.with_columns((pl.col(value_col) / base).alias(value_col))


def plot_combined_skill_premium(sp_agg: pl.DataFrame, sp_ind: pl.DataFrame, base_year: int, name_map: dict, n_highlight: int = 4) -> None:
    # Normalize aggregate at base_year.
    sp_agg_n = _normalize_series(sp_agg, "SKILL_PREMIUM", None, base_year)
    # For industry: only keep industries with a value at base_year (otherwise can't normalize).
    have_base = sp_ind.filter(pl.col("YEAR") == base_year).select("BEA_CODE")["BEA_CODE"].unique().to_list()
    sp_ind = sp_ind.filter(pl.col("BEA_CODE").is_in(have_base))
    sp_ind_n = _normalize_series(sp_ind, "SKILL_PREMIUM", "BEA_CODE", base_year)

    # Rank by ending growth.
    last_year = sp_ind_n["YEAR"].max()
    rank = sp_ind_n.filter(pl.col("YEAR") == last_year).sort("SKILL_PREMIUM")
    bottom = rank.head(n_highlight)["BEA_CODE"].to_list()
    top = rank.tail(n_highlight)["BEA_CODE"].to_list()
    highlight = set(bottom + top)

    fig, ax = plt.subplots(figsize=(11, 6.5))

    # All other industries in light grey
    for code_tuple, group in sp_ind_n.group_by("BEA_CODE"):
        code = code_tuple[0] if isinstance(code_tuple, tuple) else code_tuple
        if code in highlight:
            continue
        g = group.sort("YEAR")
        ax.plot(g["YEAR"].to_list(), g["SKILL_PREMIUM"].to_list(),
                color="#bbbbbb", linewidth=0.5, alpha=0.55, zorder=1)

    warm = ["#c0392b", "#e67e22", "#d35400", "#922b21"]
    cool = ["#1f3a5f", "#2874a6", "#117a65", "#0e6655"]
    for i, code in enumerate(reversed(top)):
        g = sp_ind_n.filter(pl.col("BEA_CODE") == code).sort("YEAR")
        last_v = g["SKILL_PREMIUM"][-1]
        label = f"{code} — {name_map.get(code, '?')[:32]} ({last_v:.2f})"
        ax.plot(g["YEAR"].to_list(), g["SKILL_PREMIUM"].to_list(),
                color=warm[i % len(warm)], linewidth=1.6, label=label, zorder=3)
    for i, code in enumerate(bottom):
        g = sp_ind_n.filter(pl.col("BEA_CODE") == code).sort("YEAR")
        last_v = g["SKILL_PREMIUM"][-1]
        label = f"{code} — {name_map.get(code, '?')[:32]} ({last_v:.2f})"
        ax.plot(g["YEAR"].to_list(), g["SKILL_PREMIUM"].to_list(),
                color=cool[i % len(cool)], linewidth=1.6, label=label, zorder=3)

    # Aggregate
    ax.plot(sp_agg_n["YEAR"].to_list(), sp_agg_n["SKILL_PREMIUM"].to_list(),
            color="black", linewidth=2.6, linestyle="--",
            label=f"Aggregate ({sp_agg_n['SKILL_PREMIUM'][-1]:.2f})", zorder=4)

    ax.axhline(1.0, color="#cccccc", linewidth=0.5, zorder=0)
    ax.set_xlim(base_year - 0.5, last_year + 0.5)
    ax.set_xlabel("Year", fontsize=11)
    ax.set_ylabel(f"Skill premium (normalized, {base_year} = 1.0)", fontsize=11)
    ax.set_title(
        f"Skill premium {base_year}–{last_year}, normalized\n"
        f"Aggregate (black dashed) + {sp_ind_n['BEA_CODE'].n_unique()} industry overlays; "
        f"top/bottom {n_highlight} by {last_year} growth",
        fontsize=12, pad=12,
    )
    style_axes(ax)
    ax.legend(loc="upper left", fontsize=8.5, frameon=False, ncol=1)
    fig.tight_layout()
    out = OUT / "fig_skill_premium_normalized.pdf"
    fig.savefig(out, bbox_inches="tight")
    fig.savefig(out.with_suffix(".png"), bbox_inches="tight", dpi=110)
    print(f"  wrote {out} (+ .png)")
    plt.close(fig)


def plot_equipment_prices_normalized(ind: pl.DataFrame, agg: pl.DataFrame, base_year: int, name_map: dict, n_highlight: int = 4) -> None:
    last_year = ind["YEAR"].max()
    rank = ind.filter(pl.col("YEAR") == last_year).sort("P_norm")
    bottom = rank.head(n_highlight)["BEA_CODE"].to_list()
    top = rank.tail(n_highlight)["BEA_CODE"].to_list()
    highlight = set(bottom + top)

    fig, ax = plt.subplots(figsize=(11, 6.5))
    for code_tuple, group in ind.group_by("BEA_CODE"):
        code = code_tuple[0] if isinstance(code_tuple, tuple) else code_tuple
        if code in highlight:
            continue
        g = group.sort("YEAR")
        ax.plot(g["YEAR"].to_list(), g["P_norm"].to_list(),
                color="#bbbbbb", linewidth=0.5, alpha=0.55, zorder=1)

    warm = ["#c0392b", "#e67e22", "#d35400", "#922b21"]
    cool = ["#1f3a5f", "#2874a6", "#117a65", "#0e6655"]
    for i, code in enumerate(reversed(top)):
        g = ind.filter(pl.col("BEA_CODE") == code).sort("YEAR")
        last_v = g["P_norm"][-1]
        ax.plot(g["YEAR"].to_list(), g["P_norm"].to_list(),
                color=warm[i % len(warm)], linewidth=1.7,
                label=f"{code} — {name_map.get(code, '?')[:32]} ({last_v:.2f})", zorder=3)
    for i, code in enumerate(bottom):
        g = ind.filter(pl.col("BEA_CODE") == code).sort("YEAR")
        last_v = g["P_norm"][-1]
        ax.plot(g["YEAR"].to_list(), g["P_norm"].to_list(),
                color=cool[i % len(cool)], linewidth=1.7,
                label=f"{code} — {name_map.get(code, '?')[:32]} ({last_v:.2f})", zorder=3)

    # Aggregate
    ag = agg.filter(pl.col("YEAR") >= base_year).sort("YEAR")
    ax.plot(ag["YEAR"].to_list(), ag["P_agg_norm"].to_list(),
            color="black", linewidth=2.6, linestyle="--",
            label=f"Aggregate ({ag['P_agg_norm'][-1]:.2f})", zorder=4)

    ax.axhline(1.0, color="#cccccc", linewidth=0.5, zorder=0)
    ax.set_xlim(base_year - 0.5, last_year + 0.5)
    ax.set_xlabel("Year", fontsize=11)
    ax.set_ylabel(f"Equipment price (nominal, {base_year} = 1.0)", fontsize=11)
    ax.set_title(
        f"Industry equipment investment prices {base_year}–{last_year}, normalized\n"
        f"All 56 industries (grey); aggregate (black dashed); top/bottom {n_highlight} by {last_year} level",
        fontsize=12, pad=12,
    )
    style_axes(ax)
    ax.legend(loc="upper right", fontsize=8.5, frameon=False, ncol=1)
    fig.tight_layout()
    out = OUT / "fig_industry_equipment_prices_normalized.pdf"
    fig.savefig(out, bbox_inches="tight")
    fig.savefig(out.with_suffix(".png"), bbox_inches="tight", dpi=110)
    print(f"  wrote {out} (+ .png)")
    plt.close(fig)


# ===========================================================================
# Main
# ===========================================================================

def main() -> None:
    print("=== Step 1: CPS load + filter + impute ===")
    df_cps = _load_cps_filtered()
    yr_min = df_cps["YEAR"].min()
    yr_max = df_cps["YEAR"].max()
    print(f"  CPS YEAR range after filters: {yr_min}–{yr_max}")

    print("\n=== Step 2: aggregate skill premium ===")
    sp_agg = build_skill_premium(df_cps)
    print(f"  computed {sp_agg.height} years: {sp_agg['YEAR'].min()}–{sp_agg['YEAR'].max()}")
    sp_agg.write_csv(ROOT / "data" / "proc" / "skill_premium_aggregate.csv")

    print("\n=== Step 3: industry-specific skill premium ===")
    sp_ind = build_industry_skill_premium(df_cps, min_n=50)
    print(f"  industry-year cells: {sp_ind.height} ({sp_ind['BEA_CODE'].n_unique()} industries)")
    sp_ind.write_csv(ROOT / "data" / "proc" / "skill_premium_by_industry.csv")

    # Determine earliest year with broad industry coverage (>= 40 of ~50 mappable industries).
    coverage = sp_ind.group_by("YEAR").agg(n_ind=pl.col("BEA_CODE").n_unique()).sort("YEAR")
    threshold = 40
    eligible = coverage.filter(pl.col("n_ind") >= threshold)
    if eligible.is_empty():
        base_year = sp_ind["YEAR"].min()
    else:
        base_year = eligible["YEAR"][0]
    print(f"  base year (>= {threshold} industries with min_n=50): {base_year}")
    print(f"  coverage summary (first 8): {coverage.head(8)}")

    names_df = pl.read_csv(INDUSTRY_NAMES)
    name_map = dict(zip(names_df["BEA_Code"].to_list(), names_df["Description"].to_list()))

    print("\n=== Step 4: plot combined skill premium ===")
    plot_combined_skill_premium(sp_agg, sp_ind, base_year, name_map)

    print("\n=== Step 5: build & plot equipment prices (full range, same base year) ===")
    ind_p, agg_p = build_industry_eq_prices(base_year)
    plot_equipment_prices_normalized(ind_p, agg_p, base_year, name_map)


if __name__ == "__main__":
    main()
